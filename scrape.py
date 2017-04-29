## Scrape from Excel data tables finding intervals where string appears, in order to lookup postcode and data
# Note: openpyxl is slower than straightout pandas (it probably could be faster, especially in ReadOnly reading time is lower,
#       especially one has to be careful setting up boundaries) however it has a better correspondence with the original Worksheet,
#       keeping empty cells and registering them as None instead of NaN.
# Going from Openpyxl to Pandas pd.DataFrame(worksheet) is the slowest procedure, should NOT be done on the whole sheet,
# even though is easiest

#Find Strings in Cells
def findStrInCell(stri,ws):
    for row in ws.iter_rows():
        for cell in row:
            if stri in str(cell.value):
               yield (cell.value, cell.row, cell.column)


#Input an array of integers and return ranges
def ranges(a):
    b = sorted(a)
    i = 0
    for j in xrange(1,len(b)):
        if b[j] > b[j-1]+1:
            yield (b[i],b[j-1])
            i = j
    yield (b[i], b[-1])

#- search ranges where string StrFnd appears in workbook flnm.
#- Input : flnm  : filename of workbook, excel or equivalent, to be read by openpyxl
#-         strFnd: string to be found in workbook, to be noted that it looks only in the last sheet.
#-         imp   : (Optional) if set to something, will not import openpyxl, for faster use if openpyxl is already imported
#- Output: list of interval ranges where the string has been found.
def search_Ranges_WorkBk(flnm,strFnd, imp = None):
    if imp is None:
        import openpyxl as op

    workb = op.load_workbook(flnm, read_only = True, data_only = True)

    worksheet = workb.worksheets[-1] #In all workbooks is the last sheet that contains data

    fndLst=list(findStrInCell(strFnd,worksheet) )
    rows=[]
    for elem in fndLst: rows.append(elem[1])

    return list(ranges(rows))

def from_Rng_to_DataFrame(flnm,rng, imp = None):
    if imp is None:
        import openpyxl as op
        import pandas as pd

    workb = op.load_workbook(flnm, data_only = True)
    works = workb.worksheets[-1] #In all workbooks is the last sheet that contains data

    headline = works[2] # Headers for these .xlsx

    headlist = []
    for cell in headline:
        headlist.append(cell.value)

    df = pd.DataFrame()

    for ranges in rng:
        for line in works.iter_rows( min_row=ranges[0], max_row=ranges[1] ):
            d = []
            for cell in line:
                d.append(cell.value)
            df = df.append(pd.Series(d),ignore_index=True)

    df.columns = headlist
    print df

# rangesList = search_Ranges_WorkBk('Pers_Mortage_PCS_2016-q3.xlsx',"London")
# print rangesList
